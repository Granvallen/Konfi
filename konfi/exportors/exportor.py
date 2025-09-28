
from pathlib import Path

import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

import json
import hashlib

from ..parsers import *
from ..writers import *
from ..etypes import EEnum


class Exportor:
    """ 导出器, 控制导出流程
    """
    
    def __init__(
        self, 
        table_dir: str, 
        data_dir: str="data",
        writer_ext: str="py",
        enum_tables: set[str]=None,
        is_inc=False,
    ) -> None:
        
        self._table_dir = table_dir # 表格文件目录
        self._data_dir = data_dir # 导出配置数据目录
        self._enum_tables = enum_tables # 枚举表集合
        self._is_inc = is_inc # 增量导表

        self._enum_table_paths: list[Path] = []
        self._config_table_paths: list[Path] = []
        self._data_path: Path = Path(self._data_dir)

        self._matchers: list[Parser] = []
        self._filters: list[Parser] = []
        self._default_parser: Parser = Parser.get_parser("CommonParser")
        self._writer: Writer = Writer.get_writer(writer_ext)
        
        self._export_data: dict[str, dict] = {} # 导出数据
        self._enum_data: dict[str, EEnum] = {} # 枚举类型
        self._md5: dict[str, str] = {}
        
        self._load_parsers()
    
    def _load_parsers(self):
        """ 加载所有解析器
        """
        parsers = Parser.get_reg_parsers()
        for name, parser in parsers.items():
            cls = type(parser)
            if "matcher" in cls.__dict__:
                self._matchers.append(parser)

            if "filter" in cls.__dict__:
                self._filters.append(parser)


    def _get_parser_for(self, sheet_name: str, ws: Worksheet) -> Parser:
        """ 获取可用的解析器
        """
        
        # 1. 匹配解析器
        for parser in self._matchers:
            if sheet_name in parser.matcher:
                return parser

        # 2. 过滤解析器
        for parser in self._filters:
            if parser.filter(ws):
                return parser

        # 3. 使用默认解析器
        return self._default_parser


    def _gather_all_tables(self):
        """ 搜索所有表格文件
        """
        path = Path(self._table_dir)
        print(f"[konfi] 搜索表格路径: {path}\n")
        
        is_valid = lambda p: p.is_file() and not p.name.startswith(("~$", "#"))
        for p in path.glob("**/*.xlsx"):
            if not is_valid(p):
                continue
            
            if self._enum_tables and (p.name in self._enum_tables or p.stem in self._enum_tables):
                self._enum_table_paths.append(p)
            else:
                self._config_table_paths.append(p)
    
    
    def _proc_mergedcell(self, ws: Worksheet):
        """ 处理合并单元格, 否则被合并的单元格值为 None
        """
        for cr in ws.merged_cells.ranges:
            min_c, min_r, max_c, max_r = cr.bounds
            top_left = ws.cell(min_r, min_c)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if r != min_r or c != min_c:
                        ws._cells[(r, c)] = top_left

    def _calc_sheet_md5(self, ws: Worksheet):
        lines = []
        for r in ws.iter_rows(values_only=True):
            line = "".join(str(v) for v in r)
            lines.append(line)
            
        md5 = hashlib.md5("\n".join(lines).encode("utf-8"))
        return md5.hexdigest()

    def _parse_sheet(self, ws: Worksheet, table_path: Path):
        """ 解析 worksheet
        """
        sheet_name = ws.title
        if sheet_name.startswith("#"): # sheet 被注释
            return 
        
        self._proc_mergedcell(ws)

        if "@" in sheet_name:
            sheet_name, _, label = sheet_name.partition("@")
        else:
            label = None

        if "." in sheet_name:
            module, _, sheet_name = sheet_name.partition(".")
        else:
            module = None
            
        if sheet_name in self._export_data:
            data = self._export_data[sheet_name]
        else:
            data = self._export_data.setdefault(sheet_name, {})
        
        # 解析 sheet
        parser = self._get_parser_for(sheet_name, ws)
        parser.parse(ws, data, self._enum_data)
        
        if info := data.get("_info"):
            info["label"] = label
            info["module"] = module
            info["sheet_type"] = parser.sheet_type
            info["table_path"] = table_path
        else:
            data["_info"] = { 
                "label": label, 
                "module": module, 
                "sheet_type": parser.sheet_type,
                "table_path": table_path,
            }

    def _parse_all_tables(self):
        """ 解析 table / workbook
        """
        # 1. 枚举表
        for table_path in self._enum_table_paths:
            wb: Workbook = xl.load_workbook(filename=table_path, read_only=False, data_only=True)
            for ws in wb.worksheets:
                self._parse_sheet(ws, table_path)

        # 2. 配置表
        for table_path in self._config_table_paths:
            wb: Workbook = xl.load_workbook(filename=table_path, read_only=False, data_only=True)
            for ws in wb.worksheets:
                if self._is_inc: # 增量导表 检测 并 更新 md5
                    ws_md5 = self._calc_sheet_md5(ws)
                    if self._md5.get(ws.title) == ws_md5:
                        continue
                    else:
                        self._md5[ws.title] = ws_md5
                
                self._parse_sheet(ws, table_path)

    
    def _write_data(self):
        """ 写入数据到文件
        """
        self._writer.write(self._export_data, self._data_path)
    
    
    def run(self):
        if self._is_inc:
            path = self._data_path / "md5.json"
            if path.is_file():
                with path.open("r", encoding="utf-8") as f:
                    self._md5 = json.load(f)
        
        # 1. 搜索所有表格文件
        self._gather_all_tables()

        # 2. 解析所有表格文件
        self._parse_all_tables()

        # 3. 导出数据
        self._write_data()

        if self._is_inc:
            print("[konfi] 更新 md5.json")
            path = self._data_path / "md5.json"
            with path.open("w", encoding="utf-8") as f:
                json.dump(self._md5, f, indent=4)



